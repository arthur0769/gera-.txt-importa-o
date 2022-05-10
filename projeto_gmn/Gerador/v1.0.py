import openpyxl
from tkinter import filedialog



local = filedialog.askopenfilename()
path = local
wb_obj = openpyxl.load_workbook(path)



planilhasA = wb_obj.worksheets
qnt_planilhas = len(planilhasA) - 1
i = 0

while i < qnt_planilhas:

    planilhasB = wb_obj.active
    qnt_linhas = planilhasB.max_row
    j = 4
    
    while j <= qnt_linhas:

        file = open(wb_obj.sheetnames[i] + ".txt", 'a')

  
        nome = wb_obj.worksheets[i].cell(row = j, column = 1)
        sexo = wb_obj.worksheets[i].cell(row = j, column = 2)
        cpf  = wb_obj.worksheets[i].cell(row = j, column = 3)
        dn   = wb_obj.worksheets[i].cell(row = j, column = 4)
        dn   = str(dn.value)
        escrevaDados = ("01|1|0000706|" + str(nome.value) + "|"+ str(sexo.value) + "|"+ str(cpf.value) + "|"+ dn +"||||||||||||||" + "\n")
        
        if ("None" in escrevaDados):
            file.close()
            break

        file.write(escrevaDados)


        coleta = wb_obj.active.cell(row = j, column = 5)
        coleta = str(coleta.value)
        escrevaColeta = ("02|2|2200|" + coleta + "\n")
        file.write(escrevaColeta)


        #exames rotina
        mensalCodigo      = "03|3|008 \n03|4|021 \n03|5|022 \n03|6|031 \n03|7|034 \n03|8|035 \n03|9|039 m \n"
        trimestralCodigo  = "03|3|008 \n03|4|021 \n03|5|022 \n03|6|031 \n03|7|034 \n03|8|035 \n03|9|039 \n03|9|099 \n03|9|099 t \n"
        semestralCodigo   = "03|3|008 \n03|4|021 \n03|5|022 \n03|6|031 \n03|7|034 \n03|8|035 \n03|9|039 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 s \n"
        anualCodigo       = "03|3|008 \n03|4|021 \n03|5|022 \n03|6|031 \n03|7|034 \n03|8|035 \n03|9|039 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 a \n"
        anual3Codigo      = "03|3|008 \n03|4|021 \n03|5|022 \n03|6|031 \n03|7|034 \n03|8|035 \n03|9|039 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 a3 \n"
        anual4Codigo      = "03|3|008 \n03|4|021 \n03|5|022 \n03|6|031 \n03|7|034 \n03|8|035 \n03|9|039 \n03|9|099 a4 \n"
        admCodigo         = "03|3|008 \n03|4|021 \n03|5|022 \n03|6|031 \n03|7|034 \n03|8|035 \n03|9|039 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 \n03|9|099 adm \n"

        mensalP     = wb_obj.worksheets[i].cell(row = j, column = 6)
        trimestralP = wb_obj.worksheets[i].cell(row = j, column = 7)
        semestralP  = wb_obj.worksheets[i].cell(row = j, column = 8)
        anualP      = wb_obj.worksheets[i].cell(row = j, column = 9)
        anual3P     = wb_obj.worksheets[i].cell(row = j, column = 10)
        anual4P     = wb_obj.worksheets[i].cell(row = j, column = 11)
        admP        = wb_obj.worksheets[i].cell(row = j, column = 12)

        if (mensalP.value != None):
            file.write(mensalCodigo)
        elif(trimestralP.value != None):
            file.write(trimestralCodigo)
        elif(semestralP.value != None):
            file.write(semestralCodigo)
        elif(anualP.value != None):
            file.write(anualCodigo)
        elif(anual3P.value != None):
            file.write(anual3Codigo)
        elif(anual4P.value != None):
            file.write(anual4Codigo)
        else:
            file.write(admCodigo)


        #exames extras

        Gcodigo     = "03|9|099 G \n"
        CREATcodigo = "03|9|099 CREAT \n"
        HGBcodigo   = "03|9|099 HGB \n"
        PROTFcodigo = "03|9|099 PROTF \n"
        FOALCcodigo = "03|9|099 FOALC \n"
        FERRIcodigo = "03|9|099 FERRI \n"
        PTHcodigo   = "03|9|099 PTH \n"
        ISTcodigo   = "03|9|099 IST \n"
        HBScodigo   = "03|9|099 HBS \n"
        HCVcodigo   = "03|9|099 HCV \n"
        VITDcodigo  = "03|9|099 VITD \n"
        HBSAGcodigo = "03|9|099 HBSAG \n"
        HIVcodigo   = "03|9|099 HIV \n"
        ALcodigo    = "03|9|099 AL \n"
        TSHcodigo   = "03|9|099 TSH \n"
        T4Lcodigo   = "03|9|099 T4L \n"
        COLFcodigo  = "03|9|099 COLF \n"

        GLICOSE     = wb_obj.worksheets[i].cell(row = j, column = 13)
        CREATININA  = wb_obj.worksheets[i].cell(row = j, column = 14)
        HGB         = wb_obj.worksheets[i].cell(row = j, column = 15)
        PROTF       = wb_obj.worksheets[i].cell(row = j, column = 16)
        FOALC       = wb_obj.worksheets[i].cell(row = j, column = 17)
        FERRITINA   = wb_obj.worksheets[i].cell(row = j, column = 18)
        PTH         = wb_obj.worksheets[i].cell(row = j, column = 19)
        IST         = wb_obj.worksheets[i].cell(row = j, column = 20)
        HBS         = wb_obj.worksheets[i].cell(row = j, column = 21)
        HCV         = wb_obj.worksheets[i].cell(row = j, column = 22)
        VITD        = wb_obj.worksheets[i].cell(row = j, column = 23)
        HBSAG       = wb_obj.worksheets[i].cell(row = j, column = 24)
        HIV         = wb_obj.worksheets[i].cell(row = j, column = 25)
        AL          = wb_obj.worksheets[i].cell(row = j, column = 26)
        TSH         = wb_obj.worksheets[i].cell(row = j, column = 27)
        T4L         = wb_obj.worksheets[i].cell(row = j, column = 28)
        COLF        = wb_obj.worksheets[i].cell(row = j, column = 29)

        if (GLICOSE.value != None):
            file.write(Gcodigo)

        if(CREATININA.value != None):
            file.write(CREATcodigo)

        if(HGB.value != None):
            file.write(HGBcodigo)

        if(PROTF.value != None):
            file.write(PROTFcodigo)

        if(FOALC.value != None):
            file.write(FOALCcodigo)

        if(FERRITINA.value != None):
            file.write(FERRIcodigo)
        
        if(PTH.value != None):
            file.write(PTHcodigo)

        if(IST.value != None):
            file.write(ISTcodigo)

        if(HBS.value != None):
            file.write(HBScodigo)

        if(HCV.value != None):
            file.write(HCVcodigo)

        if(VITD.value != None):
            file.write(VITDcodigo)

        if(HBSAG.value != None):
            file.write(HBSAGcodigo)

        if(HIV.value != None):
            file.write(HIVcodigo)

        if(AL.value != None):
            file.write(ALcodigo)

        if(TSH.value != None):
            file.write(TSHcodigo)

        if(T4L.value != None):
            file.write(T4Lcodigo)

        if(COLF.value != None):
            file.write(COLFcodigo)
        
        
        


        file.close()

            

            
            

            


        j+=1

    i+=1